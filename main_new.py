#!/usr/bin/python
# -*- coding: utf-8 -*-
import os

import pyodbc
import openpyxl
import notification
from openpyxl import Workbook
import mailsender
from datetime import date
import time
import getpass
import sys


class PorscheLeasing:
    wb = Workbook()
    current_dir = "c:\\AM\\BAT\\porsche_leasing\\"

    wb.save(current_dir + 'work_and_spare_parts.xlsx')

    ##########connection to sql################
    filename = current_dir + 'work_and_spare_parts.xlsx'
    server = ''
    database = ''
    username = ''
    password = ''
    driver = '{SQL Server}'  # Driver you need to connect to the database
    port = '1433'

    def __init__(self, gsalid):
        self.gsalid = gsalid
        self.cnn = pyodbc.connect(
            'DRIVER=' + self.driver + ';PORT=port;SERVER=' + self.server + ';PORT=1443;DATABASE=' + self.database + ';UID=' + self.username +
            ';PWD=' + self.password)
        self.cursor = self.cnn.cursor()

    def connection_to_base(self):
        try:
            sql_query_1 = f"""
            declare @gsalid int
            set @gsalid={self.gsalid}
            select '24200' as dealercode,'320984' as mfo,' UA333209840000026008210383131' as rr,s.name,convert(varchar(10),g.wrkordno)+'/'+convert(varchar(10),b.grecno) as grecno,g.created,v.SERIALNO,g.DISTDRIV,cast ((cast(m.rsum as  numeric(36,2))*1.2) as  numeric(36,2))   from GSALS01 g
             join sman_full s on s.smanid=g.RELINQUI
            join GBILS01 b on b.GSALID=g.GSALID
            join vehi v on v.vehiid=g.VEHIID
             join (select sum(rsum/1.2) as rsum,GSALID from GROWS01
              where gsalid=@gsalid
              group by gsalid)  m on m.GSALID=g.GSALID
             where g.gsalid= @gsalid


            """

            res = self.cursor.execute(sql_query_1)
            res_list = list()
            for item in res:
                res_list.append(item)

            wb = openpyxl.load_workbook(self.filename)
            w1 = wb["Sheet"]
            w1.cell(row=1, column=1).value = "Код дилера"
            w1.cell(row=2, column=1).value = "МФО банку"
            w1.cell(row=3, column=1).value = "Номер рахунку в банку"
            w1.cell(row=4, column=1).value = "ПІБ майстра-приймальника"
            w1.cell(row=5, column=1).value = "№ документу"
            w1.cell(row=6, column=1).value = "Дата документу"
            w1.cell(row=7, column=1).value = "VIN код автомобіля"
            w1.cell(row=8, column=1).value = "Пробіг автомобіля"
            w1.cell(row=9, column=1).value = "Всього з ПДВ"
            w1.cell(row=10, column=1).value = ""
            w1.cell(row=1, column=2).value = res_list[0][0]
            w1.cell(row=2, column=2).value = res_list[0][1]
            w1.cell(row=3, column=2).value = res_list[0][2]
            w1.cell(row=4, column=2).value = res_list[0][3]
            w1.cell(row=5, column=2).value = res_list[0][4]
            try:
                w1.cell(row=6, column=2).value = (res_list[0][5].strftime("%d.%m.%Y"))
            except Exception as err:
                print(err)
                w1.cell(row=6, column=2).value = ""
            w1.cell(row=7, column=2).value = res_list[0][6]
            w1.cell(row=8, column=2).value = res_list[0][7]
            w1.cell(row=9, column=2).value = float(res_list[0][8])
            w1.cell(row=10, column=2).value = ""
            w1.column_dimensions['A'].width = 30
            w1.column_dimensions['B'].width = 30
            w1.column_dimensions['C'].width = 30
            w1.column_dimensions['D'].width = 30
            w1.column_dimensions['E'].width = 30
            w1.column_dimensions['F'].width = 30
            w1.column_dimensions['G'].width = 30
            w1.column_dimensions['H'].width = 30
            number_doc_without_slash = res_list[0][4].replace("/", "_")
            print(number_doc_without_slash)
            wb.save(self.filename)
            abs_path_with_filename = "{}\{}.xlsx".format(self.current_dir, number_doc_without_slash)
            try:
                os.rename("{}\work_and_spare_parts.xlsx".format(self.current_dir), abs_path_with_filename)
            except Exception as err:
                os.remove(abs_path_with_filename)
                os.rename("{}\work_and_spare_parts.xlsx".format(self.current_dir), abs_path_with_filename)

            self.filename = abs_path_with_filename
        except Exception as err:
            print(err)

    def connection_to_base2(self):
        try:

            sql_query_2 = f"""declare @gsalid int

            set @gsalid={self.gsalid}


            select number=row_number() over(order by  g.rtype ),case when g.rtype=1 then 'запчастина' when g.rtype=7 then 'робота' end as роботазапчастина,g.item,cast(g.num as numeric(36,2)) as num ,case when rtype=1 then i.SpSiZEUNIT
            when rtype=7 then '|'end  as category,g.name,cast(round((g.rsum/1.2),2) as numeric(36,2)) as sum from GROWS01 g
            left join item i on i.ITEMNO=g.itemno and i.suplno=g.suplno and g.rtype=1

            where gsalid=@gsalid and rtype in (1,7)
            order by g.rtype

            """

            res = self.cursor.execute(sql_query_2)
            wb = openpyxl.load_workbook(self.filename)
            w1 = wb["Sheet"]
            w1.cell(row=10, column=1).value = "# пп"
            w1.cell(row=10, column=2).value = "робота/запчастина"
            w1.cell(row=10, column=3).value = "код (роботи/запчастини)"
            w1.cell(row=10, column=4).value = "кількість (нормогодин/запчастин)"
            w1.cell(row=10, column=5).value = "категорія роботи / одиниця виміру запчастин"
            w1.cell(row=10, column=6).value = "назва"
            w1.cell(row=10, column=7).value = "вартість зі знижкою без ПДВ"

            for item in res:
                w1.append(list(item))

            wb.save(self.filename)

        except Exception as err:
            print(err)

    def add_to_table(self, author):
        try:
            sql_query3 = """insert into amintegrations.dbo.pl (date, gsalid, author)
    values (getdate(), '{}', '{}')""".format(self.gsalid, author)
            self.cursor.execute(sql_query3)
            self.cnn.commit()  # заполняем колонки в базе если соблюдается условие
        except Exception as err:
            print(err)


if __name__ == '__main__':
    #    gsalid = sys.argv[1]
    today = date.today()
    gsalid = 233984
    main_prog = PorscheLeasing(gsalid)#sys.argv[1])
    main_prog.connection_to_base()
    main_prog.connection_to_base2()
    main_prog.add_to_table(getpass.getuser())
    # add attachment with emeils test@test.ua;test2@test.ua and filename
    mailsender.add_attachments("o.koval@avtosojuz.ua",
                               main_prog.filename)
    notification.my_notifier()
    os.remove(main_prog.filename)
    main_prog.cursor.close()
    main_prog.cnn.close()
