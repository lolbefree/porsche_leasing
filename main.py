#!/usr/bin/python
# -*- coding: utf-8 -*-
import os

import pyodbc
import openpyxl
import notification
from openpyxl import Workbook
import mailsender
from datetime import date
import time
import getpass

wb = Workbook()
current_dir = os.path.abspath(os.curdir)
wb.save(filename=current_dir+'work_and_spare_parts.xlsx')

##########connection to sql################
filename='work_and_spare_parts.xlsx'
server = ''
database = ''
username = ''
password = ''
driver = '{SQL Server}'  # Driver you need to connect to the database
port = '1433'


def connection_to_base(gsalid):
    global filename, current_dir
    try:
        sql_query_1 = f"""
        declare @gsalid int
        set @gsalid={gsalid}
        select '24200' as dealercode,'320984' as mfo,' UA333209840000026008210383131' as rr,s.name,convert(varchar(10),g.wrkordno)+'/'+convert(varchar(10),b.grecno) as grecno,b.billd,v.SERIALNO,g.DISTDRIV,cast ((cast(m.rsum as  numeric(36,2))*1.2) as  numeric(36,2))   from GSALS01 g
         join sman_full s on s.smanid=g.RELINQUI
        join GBILS01 b on b.GSALID=g.GSALID
        join vehi v on v.vehiid=g.VEHIID
         join (select sum(rsum/1.2) as rsum,GSALID from GROWS01
          where gsalid=@gsalid
          group by gsalid)  m on m.GSALID=g.GSALID
         where g.gsalid= @gsalid


        """

        cnn = pyodbc.connect(
            'DRIVER=' + driver + ';PORT=port;SERVER=' + server + ';PORT=1443;DATABASE=' + database + ';UID=' + username +
            ';PWD=' + password)
        cursor = cnn.cursor()
        res = cursor.execute(sql_query_1)
        res_list = []
        for item in res:
            res_list.append(item)

        wb = openpyxl.load_workbook(filename)
        w1 = wb["Sheet"]
        w1.cell(row=1, column=1).value = "Код дилера"
        w1.cell(row=2, column=1).value = "МФО банку"
        w1.cell(row=3, column=1).value = "Номер рахунку в банку"
        w1.cell(row=4, column=1).value = "ПІБ майстра-приймальника"
        w1.cell(row=5, column=1).value = "№ документу"
        w1.cell(row=6, column=1).value = "Дата документу"
        w1.cell(row=7, column=1).value = "VIN код автомобіля"
        w1.cell(row=8, column=1).value = "Пробіг автомобіля"
        w1.cell(row=9, column=1).value = "Всього з ПДВ"
        w1.cell(row=10, column=1).value = ""
        w1.cell(row=1, column=2).value = res_list[0][0]
        w1.cell(row=2, column=2).value = res_list[0][1]
        w1.cell(row=3, column=2).value = res_list[0][2]
        w1.cell(row=4, column=2).value = res_list[0][3]
        w1.cell(row=5, column=2).value = res_list[0][4]
        try:
            w1.cell(row=6, column=2).value = (res_list[0][5].strftime("%d.%m.%Y"))
        except Exception as err:
            print(err)
            w1.cell(row=6, column=2).value = ""
        w1.cell(row=7, column=2).value = res_list[0][6]
        w1.cell(row=8, column=2).value = res_list[0][7]
        w1.cell(row=9, column=2).value = float(res_list[0][8])
        w1.cell(row=10, column=2).value = ""
        w1.column_dimensions['A'].width = 30
        w1.column_dimensions['B'].width = 30
        w1.column_dimensions['C'].width = 30
        w1.column_dimensions['D'].width = 30
        w1.column_dimensions['E'].width = 30
        w1.column_dimensions['F'].width = 30
        w1.column_dimensions['G'].width = 30
        w1.column_dimensions['H'].width = 30

        cursor.close()
        cnn.close()
        number_doc_without_slash = res_list[0][4].replace("/", "_")
        print(number_doc_without_slash)
        wb.save(filename)
        abs_path_with_filename ="{}\{}.xlsx".format(current_dir, number_doc_without_slash)
        try:
            os.rename("{}\work_and_spare_parts.xlsx".format(current_dir), abs_path_with_filename)
        except Exception as err:
            os.remove(abs_path_with_filename)
            os.rename("{}\work_and_spare_parts.xlsx".format(current_dir), abs_path_with_filename)

        filename = abs_path_with_filename
    except Exception as err:
        print(err)
print("tyt", filename)

def connection_to_base2(gsalid):
    global filename
    try:

        sql_query_2 = f"""declare @gsalid int

        set @gsalid={gsalid}


        select number=row_number() over(order by  g.rtype ),case when g.rtype=1 then 'запчастина' when g.rtype=7 then 'робота' end as роботазапчастина,g.item,cast(g.num as numeric(36,2)) as num ,case when rtype=1 then i.SpSiZEUNIT
        when rtype=7 then '|'end  as category,g.name,cast(round((g.rsum/1.2),2) as numeric(36,2)) as sum from GROWS01 g
        left join item i on i.ITEMNO=g.itemno and i.suplno=g.suplno and g.rtype=1

        where gsalid=@gsalid and rtype in (1,7)
        order by g.rtype

        """

        cnn = pyodbc.connect(
            'DRIVER=' + driver + ';PORT=port;SERVER=' + server + ';PORT=1443;DATABASE=' + database + ';UID=' + username +
            ';PWD=' + password)
        cursor = cnn.cursor()
        res = cursor.execute(sql_query_2)

        wb = openpyxl.load_workbook(filename)
        w1 = wb["Sheet"]
        w1.cell(row=10, column=1).value = "# пп"
        w1.cell(row=10, column=2).value = "робота/запчастина"
        w1.cell(row=10, column=3).value = "код (роботи/запчастини)"
        w1.cell(row=10, column=4).value = "кількість (нормогодин/запчастин)"
        w1.cell(row=10, column=5).value = "категорія роботи / одиниця виміру запчастин"
        w1.cell(row=10, column=6).value = "назва"
        w1.cell(row=10, column=7).value = "вартість зі знижкою без ПДВ"

        for item in res:
            w1.append(list(item))

        cursor.close()
        cnn.close()
        wb.save(filename)

    except Exception as err:
        print(err)


def add_to_table(gsalid, author):
    try:
        sql_query3 = """insert into amintegrations.dbo.pl (date, gsalid, author)
values (getdate(), '{}', '{}')""".format(gsalid, author)
        cnn = pyodbc.connect(
            'DRIVER=' + driver + ';PORT=port;SERVER=' + server + ';PORT=1443;DATABASE=' + database + ';UID=' + username +
            ';PWD=' + password)
        cursor = cnn.cursor()
        cursor.execute(sql_query3)
        cnn.commit()  # заполняем колонки в базе если соблюдается условие

        cursor.close()
        cnn.close()
    except Exception as err:
        print(err)


if __name__ == '__main__':
    # gsalid = sys.argv[1]
    today = date.today()
    gsalid = 271295

    login = getpass.getuser()
    connection_to_base(gsalid)
    connection_to_base2(gsalid)
    wb = openpyxl.load_workbook(filename)
    w1 = wb["Sheet"]
    w1.insert_rows(10)
    wb.save(filename)
    # add attachment with emeils test@test.ua;test2@test.ua and filename
    mailsender.add_attachments("o.koval@avtosojuz.ua", filename)
   # mailsender.sender("oleh.shcherban@porschefinance.ua;oleksandr.pievien@porschefinance.ua", "", name_file)
    add_to_table(int(gsalid), str(login))
    notification.my_notifier()
    os.remove(filename)
# oleh.shcherban@porschefinance.ua
# oleksandr.pievien@porschefinance.ua
#oleh.shcherban@porschefinance.ua;oleksandr.pievien@porschefinance.ua
#o.koval@avtosojuz.ua
#d@avtosojuz.ua




